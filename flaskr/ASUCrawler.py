import settings
from variables import Globals
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, WebDriverException
import time
import requests
import xlsxwriter
import datetime as dt
from bs4 import BeautifulSoup
import sys, os
import openpyxl
import re

class Crawler:
	global website
	filename = ''
	tabName = ''
	opt = ''
	date = dt.datetime.today().strftime("%m-%d-%Y")
	file = open('logs.txt','a')
	global workbook
	'''def spiderCrawler(self,site,username,password):
		s = requests.Session()
		file = open("output.txt",'a')
		file2 = open("html.txt",'a')
		headers = {
		'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; WOW64; rv:27.0) Gecko/20100101 Firefox/27.0'
		}
		loginRequest = s.get('https://webapp4.asu.edu/myasu',headers=headers,verify=False)
		responseCookies = loginRequest.cookies
		# print responseCookies
		cookies = requests.utils.cookiejar_from_dict(requests.utils.dict_from_cookiejar(s.cookies))
		print cookies
		soup = BeautifulSoup(loginRequest.text,'html.parser')
		lt = soup.find("input",{"name":"lt"})
		execution = soup.find("input",{"name":"execution"})
		event = soup.find("input",{"name":"_eventId"})
		ltValue = lt['value']
		executionValue = execution['value']
		eventValue = event['value']
		print ltValue
		print executionValue
		print eventValue
		print username
		print password
		data = {
		"username" : username,
		"password" : password,
		"lt" : ltValue,
		"execution" : executionValue,
		"_eventId" : eventValue
		}
		headers = {
		'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; WOW64; rv:27.0) Gecko/20100101 Firefox/27.0',
		'Referer:' : '"https://weblogin.asu.edu/cas/login?service=https%3A%2F%2Fweblogin.asu.edu%2Fcgi-bin%2Fcas-login%3Fcallapp%3Dhttps%253A%252F%252Fwebapp4.asu.edu%252Fmyasu%252F%253Finit%253Dfalse"'
		}
		loginRequest = s.post('https://weblogin.asu.edu/cas/login?service=https%3A%2F%2Fweblogin.asu.edu%2Fcgi-bin%2Fcas-login%3Fcallapp%3Dhttps%253A%252F%252Fwebapp4.asu.edu%252Fmyasu%252F%253Finit%253Dfalse',data=data,cookies=cookies,headers=headers)
		print loginRequest.status_code
		headers = {
		'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; WOW64; rv:27.0) Gecko/20100101 Firefox/27.0'
		}
		success = s.get('https://webapp4.asu.edu/myasu/')
		file2.write(success.text)
		file2.close()
		captureLinks = BeautifulSoup(success.text,'html.parser')
		href_tags = captureLinks.find_all(href=True)
		# for link in captureLinks.find_all('a',href=True):
			# file.write("%s\n" % link['href'])
		for link in href_tags:
			file.write("%s\n" % link['href'])
		file.close()'''


	def goToAllLinks(self,site):
		print Crawler.opt
		print Crawler.date
		Crawler.workbook = xlsxwriter.Workbook(Crawler.date+"_"+Crawler.opt+".xlsx")
		links = ['student/finances','student/campusservices','student/profile',
				'staff/resources','staff/service','staff/profile','student']
		addLater = ['student/finances']
		for link in links:
			linkSplit = link.split('/')
			Crawler.tabName = link
			currentLink = site+link
			print currentLink
			self.captureLinks(currentLink)
		print "Returned to goToAllLinks"
		self.compareWithBaseline()
		print "Exiting."

	def captureLinks(self,site):
		global linkSet
		# declare set to store unique links
		linkSet = set()
		print "Currently crawling "+site
		# load myASU
		Globals.driver.get(site)
		# WebDriverWait(Globals.driver,10).until(self.readystate() == True)
		try:
			# wait for page to load
			'''WebDriverWait(Globals.driver,15).until(
				EC.presence_of_element_located((By.ID,"asu_footer"))
				)'''
			WebDriverWait(Globals.driver,15).until(self.ajax_complete,"Timeout waiting for Ajax")
			# Globals.driver.implicitly_wait(15)
		except TimeoutException:
			# timeout
			print "Timeout"
		# capture all anchor tags on the page
		links = Globals.driver.find_elements_by_css_selector('a[href]')
		print str(len(links))
		# iterate over all the links on the page
		for link in links:
			try:
				href = link.get_attribute("href")
				if href:
					# remove invalid links
					if "myasu/Signout" not in href and "myasu/pscs" not in href and "javascript" not in href and "keep" not in href and "@asu.edu" not in href:
						if href not in linkSet:
							print href
							# add valid links to the set
							linkSet.add(href)
			except Exception as e:
				print "Exception!"
				continue
		print "All links added to the set"
		self.writeToFile(site,linkSet)

	def writeToFile(self,site,linkSet):
		print "I'm here"
		Crawler.tabName = Crawler.tabName.replace("/","-")
		worksheet = Crawler.workbook.add_worksheet(Crawler.tabName)
		row = 0
		col = 0
		for link in linkSet:
			try:
				Globals.driver.get(link)
				Globals.driver.implicitly_wait(5)
			except Exception as e:
				Crawler.file.write("%s\n" % "Error occured for link "+link)
				print "Error occured for link "+link
				print sys.exc_info()[0]
				continue
			windowTitle = Globals.driver.title
			worksheet.write(row,col,windowTitle)
			worksheet.write(row,col+1,link)
			row += 1
		print "function writeToFile executed"
		# Crawler.workbook.close()
		# file.close()

	def compareWithBaseline(self):
		Crawler.workbook.close()
		workbookBaseline = openpyxl.load_workbook(os.path.join(settings.app.config['UPLOAD_FOLDER'], Crawler.filename))
		workbookCurrent = openpyxl.load_workbook(Crawler.date+"_"+Crawler.opt+".xlsx")
		name = os.path.join(settings.app.config['UPLOAD_FOLDER'], Crawler.filename)
		print name
		print Crawler.date
		print "Loaded baseline file"
		newLinksFile = open('newLinks_'+Crawler.opt+'_'+Crawler.date+'.txt','a')
		removedLinksFile = open('removedLinks_'+Crawler.opt+'_'+Crawler.date+'.txt','a')
		for (sheetBaseline,sheetCurrent) in zip(workbookBaseline,workbookCurrent):
			baselineSet = set()
			currentSet = set()
			print sheetBaseline
			print sheetCurrent
			rowsBaseline = sheetBaseline.get_highest_row()
			rowsCurrent = sheetCurrent.get_highest_row()
			print "Entering for loop for sheet iteration"
			for row in range(1,rowsBaseline):
				print "Created set for baseline"
				cell_name = "{}{}".format("B",row)
				print cell_name
				currentString = sheetBaseline[cell_name].value
				if "-qa" in currentString:
					# currentString = re.sub('https://[^/]+/','',currentString)
					tempString = currentString.replace("-qa","")
					baselineSet.add(tempString)
				else:	
					baselineSet.add(currentString)
			for row in range(1,rowsCurrent):
				print "Created set for current link"
				cell_name = "{}{}".format("B",row)
				currentString = sheetCurrent[cell_name].value
				if "-qa" in currentString:
					# currentString = re.sub('https://[^/]+/','',currentString)
					tempString = currentString.replace("-qa","")
					currentSet.add(tempString)
				else:
					currentSet.add(currentString)
			removedLinksList = list(baselineSet - currentSet)
			print "Removed links list created"
			for deletedLink in removedLinksList:
				print deletedLink
			addedLinksList = list(currentSet - baselineSet)
			for newLink in addedLinksList:
				print newLink	
			print "Added links list created"
			removedLinksFile.write("\n%s\n" % "Tab: "+sheetBaseline.title)
			for entry in removedLinksList:
				removedLinksFile.write("%s\n" % ""+entry)
			newLinksFile.write("%s\n" % "Tab: "+sheetBaseline.title)
			for entry in addedLinksList:
				newLinksFile.write("%s\n" % ""+entry)
		print "Closing all files"
		newLinksFile.close()
		removedLinksFile.close()
		file.close()


	def ajax_complete(self,driver):
	    try:
	        return 0 == driver.execute_script("return jQuery.active")
	    except WebDriverException:
	        pass

	def readystate(self):
		result = Globals.driver.execute_script("return document.readyState;")
		print result
		if result == 'complete':
			flag = True
		else:
			flag = False
		print flag
		return flag

	def login(self,username,password,site,option,fileName):
		import re
		# get request to MyASU login portal
		Globals.driver.get("https://webapp4.asu.edu/myasu/")
		print option
		Crawler.filename = fileName
		print Crawler.filename
		Crawler.opt = option
		if option == "QA":
			website = "https://webapp4-qa.asu.edu/myasu/"
		elif option == "Production":
			website = "https://webapp4.asu.edu/myasu/"
		else:
			wesbite = site
		try:
			# wait for sign-in button
			signIn = WebDriverWait(Globals.driver,20).until(
				EC.presence_of_element_located((By.ID,"login_submit"))
				)
			# grab username and password elements
			user = Globals.driver.find_element_by_id("username")
			passW = Globals.driver.find_element_by_id("password")
			# attempt login to myASU
			user.send_keys(username)
			passW.send_keys(password)
			Globals.driver.find_element_by_class_name("submit").click()
			WebDriverWait(Globals.driver,20).until(
				EC.presence_of_element_located((By.ID,"asu_footer"))
				)
		except TimeoutException:
			print "Exception!"
			Globals.driver.back()
			Globals.driver.back()
		finally:
			src = Globals.driver.page_source
			loggedIn = 'My ASU'
			isSignedIn = loggedIn in src
			if isSignedIn:
				print "You're good!"
				print website
				self.goToAllLinks(website)
			else:
				print "Try again"
				Globals.driver.back()
				Globals.driver.back()



